import os
import sys
from datetime import date
import pandas as pd
from datetime import datetime
if os.path.basename(os.getcwd()) == "openpolicedata":
    sys.path.append(os.path.join("..","openpolicedata"))
    output_dir = os.path.join(".","data","backup")
else:
    sys.path.append(os.path.join("..","..","openpolicedata"))
    output_dir = os.path.join("..","data","backup")
import openpolicedata as opd

run_all = True
run_all_years = True
run_all_agencies = True
log_result = True
if log_result:
    from openpyxl import Workbook, load_workbook
skip_sources = [("California","STOPS"),("Fayetteville", "ALL"),("Los Angeles County", "ALL")]
istart = 323
datasets = opd.datasets.query()
if run_all:
    max_num_stanford = float("inf")
else:
    max_num_stanford = 1

xl_dir = os.path.join(output_dir, "standardization")
if not os.path.exists(xl_dir):
    os.mkdir(xl_dir)

def autoset_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

def log_to_excel(table, year):
    table_name = table.table_type.value.replace("/","-")
    xl_filename = os.path.join(xl_dir,f"{table.state}_{table.source_name}_{table.agency}_{table_name}.xlsx")
    if os.path.exists(xl_filename):
        wb = load_workbook(xl_filename)
    else:
        wb = Workbook()
        
    cols = [x for x in table.table.columns if not x.startswith("RAW_")]
    mat = []
    for x in cols:
        loc = [k for k in range(len(table.clean_hist)) if x == table.clean_hist[k].new_column_name]
        if len(loc)==0:
            mat.append([x,None])
        elif len(loc)>1:
            raise ValueError("This should not happen")
        else:
            map = table.clean_hist[loc[0]]
            if type(map.old_column_name) == str:
                mat.append([map.old_column_name,x])
            else:
                for name in map.old_column_name:
                    mat.append([name,x])
    
    sheets = wb.sheetnames
    if sheets == ["Sheet"]:
        # Default val. Rename sheet.
        wb["Sheet"].title = "Columns"

    ws = wb["Columns"]
    if ws.cell(row = 1, column = 1).value == None:
        update = True
        year_range = str(year)
        col = 1
    else:
        col = 1
        while ws.cell(row = 1, column = col).value != None:
            col+=2
        col-=2
        mat_old = []
        row = 3
        while ws.cell(row = row, column = col).value != None:
            if ws.cell(row = row, column = col+1).value==None:
                mat_old.append([ws.cell(row = row, column = col).value,None])
            else:
                mat_old.append([ws.cell(row = row, column = col).value,ws.cell(row = row, column = col+1).value])
            row+=1

        update = mat != mat_old
        if update:
            year_range = str(year)
            col+=2
        else:
            if "-" in ws.cell(row = 1, column = col).value:
                year_range = [int(x) for x in ws.cell(row = 1, column = col).value.split("-")]
            else:
                year_range = int(ws.cell(row = 1, column = col).value)
                year_range = [year_range,year_range]
            if year == year_range[0]-1:
                year_range[0]-=1
            else:
                year_range[0] = year

            year_range = f"{year_range[0]}-{year_range[1]}"
    ws.cell(row = 1, column = col).value = str(year_range)

    if update:
        ws.cell(row = 2, column = col).value = "Original"
        ws.cell(row = 2, column = col+1).value = "Cleaned"
        row = 2

        for k,row in enumerate(mat):
            ws.cell(row = 3+k, column = col).value = row[0]
            if row[1] != None:
                ws.cell(row = 3+k, column = col+1).value = row[1]

        autoset_width(ws)

    for map in table.clean_hist:
        if type(map.new_column_name) == str:
            if map.new_column_name not in wb.sheetnames:
                wb.create_sheet(map.new_column_name)
            ws_map = wb[map.new_column_name]
            col = 1
            while ws_map.cell(row = 1, column = col).value != None:
                col+=2
            ws_map.cell(row = 1, column = col).value = str(year)
            row = 2
            if map.data_maps != None:
                if type(map.data_maps)==dict:
                    for x,y in map.data_maps.items():
                        ws_map.cell(row = row, column = col).value = x
                        ws_map.cell(row = row, column = col+1).value = y
                        row+=1
                else:
                    for m in map.data_maps:
                        for x,y in m.items():
                            ws_map.cell(row = row, column = col).value = x
                            ws_map.cell(row = row, column = col+1).value = y
                            row+=1

            counts = table.table[map.new_column_name].value_counts()
            for k in range(min(10,len(counts))):
                ws_map.cell(row = row+k, column = col).value = str(counts.index[k])
                ws_map.cell(row = row+k, column = col+1).value = counts.iloc[k]

            autoset_width(ws_map)
        else:
            raise TypeError(f"Unknown type for new column name {map.new_column_name}")

    wb.save(xl_filename)

num_stanford = 0
prev_sources = []
prev_tables = []
prev_states = []
output_dir = ".\\data"
move_to_end = []
has_issue = datasets["SourceName"].apply(lambda x : x in move_to_end)
no_issue = datasets["SourceName"].apply(lambda x : x not in move_to_end)
datasets = pd.concat([datasets[no_issue], datasets[has_issue]])
for i in range(istart, len(datasets)):
    if "stanford.edu" in datasets.iloc[i]["URL"]:
        num_stanford += 1
        if num_stanford > max_num_stanford:
            continue

    srcName = datasets.iloc[i]["SourceName"]
    state = datasets.iloc[i]["State"]

    if not run_all_agencies and datasets.iloc[i]["Agency"] == opd.defs.MULTI and srcName == "Virginia":
        # Reduce size of data load by filtering by agency
        agency = "Fairfax County Police Department"
    else:
        agency = None

    skip = False
    for k in range(len(prev_sources)):
        if srcName == prev_sources[k] and datasets.iloc[i]["TableType"] ==prev_tables[k] and \
            datasets.iloc[i]["State"] == prev_states[k]:
            skip = True

    if skip or any([x==srcName and (y==datasets.iloc[i]["TableType"] or y=="ALL") for x,y in skip_sources]):
        continue

    prev_sources.append(srcName)
    prev_tables.append(datasets.iloc[i]["TableType"])
    prev_states.append(datasets.iloc[i]["State"])

    table_print = datasets.iloc[i]["TableType"]
    now = datetime.now().strftime("%d.%b %Y %H:%M:%S")
    print(f"{now} Running dataset {i} of {len(datasets)}: {srcName} {table_print} table")

    src = opd.Source(srcName, state=state)

    year = date.today().year
    table = None
    csv_filename = "NOT A FILE"
    try:
        years = src.get_years(datasets.iloc[i]["TableType"])
        years.sort(reverse=True)
        load_by_year = True
    except:
        load_by_year = False

    if load_by_year:
        for y in years:
            try:
                csv_filename = src.get_csv_filename(y, output_dir, datasets.iloc[i]["TableType"], 
                    agency=agency)
            except ValueError as e:
                if "There are no sources matching tableType" in e.args[0]:
                    continue
                else:
                    raise
            except:
                raise
            
            if os.path.exists(csv_filename):
                table = src.load_from_csv(y, table_type=datasets.iloc[i]["TableType"], 
                    agency=agency,output_dir=output_dir)
            elif run_all_years:
                backup_dir = os.path.join(output_dir, "backup")
                csv_filename = src.get_csv_filename(y, "", datasets.iloc[i]["TableType"], agency=agency)
                zip_filename = csv_filename.replace(".csv",".zip")
                if not os.path.exists(os.path.join(backup_dir,zip_filename)):
                    # raise FileNotFoundError(f"Backup cannot be found for file {zip_filename}")
                    table = src.load_from_url(y, table_type=datasets.iloc[i]["TableType"], 
                        agency=agency)
                else:
                    table = src.load_from_csv(y, table_type=datasets.iloc[i]["TableType"], 
                        agency=agency,output_dir=backup_dir, filename=zip_filename)
                print(f"\t{y} successfully read")
            else:
                continue
            
            table.clean(keep_raw=True)
            table.merge_date_and_time(ifmissing="ignore")

            if log_result:
                log_to_excel(table, y)
                    
            if not run_all_years:
                break
    else:
        csv_filename = src.get_csv_filename(datasets.iloc[i]["Year"], output_dir, datasets.iloc[i]["TableType"], 
                    agency=agency)
        if os.path.exists(csv_filename):
            table = src.load_from_csv(datasets.iloc[i]["Year"], table_type=datasets.iloc[i]["TableType"], 
                agency=agency,output_dir=output_dir)
        elif run_all:
            backup_dir = os.path.join(output_dir, "backup")
            csv_filename = src.get_csv_filename(datasets.iloc[i]["Year"], "", datasets.iloc[i]["TableType"], agency=agency)
            zip_filename = csv_filename.replace(".csv",".zip")
            if not os.path.exists(os.path.join(backup_dir,zip_filename)):
                # raise FileNotFoundError(f"Backup cannot be found for file {zip_filename}")
                table = src.load_from_url(datasets.iloc[i]["Year"], table_type=datasets.iloc[i]["TableType"], 
                    agency=agency)
            else:
                table = src.load_from_csv(datasets.iloc[i]["Year"], table_type=datasets.iloc[i]["TableType"], 
                    agency=agency,output_dir=backup_dir, filename=zip_filename)
        else:
            raise FileNotFoundError(f"File {csv_filename} not found")

        table.clean(keep_raw=True)
        table.merge_date_and_time(ifmissing="ignore")

        if log_result:
            log_to_excel(table, datasets.iloc[i]["Year"])
    

print("data main function complete")
